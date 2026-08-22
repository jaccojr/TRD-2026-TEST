"""
Reads camping-schema.xlsx and produces the campSchedule list used by build_data.py.

The spreadsheet's column order (Datum/tijd, Campsite, Actief?, Diner, Dagbriefing,
Ontbijt, Start, Menu, Verzorgingspost 1, Verzorgingspost 2, Bar open, Bar dicht,
Stilte, Rugzakjes afgifte, Rugzakjes ophalen, Registratie, Opmerking) is the actual
source of truth for both VALUES and DISPLAY ORDER for the original 7 columns.
The app's rendering code should read rows in this same left-to-right order rather
than maintaining its own separate field sequence — that mismatch (app used
Ontbijt/Start/Diner while the sheet always had Diner/Ontbijt/Start) is what caused
the evening-row ordering bug. Dagbriefing was added the same way after it was
found hardcoded to "20:30" directly in the app instead of coming from the sheet.

The columns after Start (Menu onward) were added for the September-2026 organizer
handoff (4-camp schedule). They're captured here so the facts are never lost, but
as of this version the app's UI only reads diner/dagbriefing/ontbijt/start/menu —
the rest (verzorgingspost, bar hours, stilte, rugzakjes, registratie, opmerking)
are recorded on each row and available in data.js for whenever that UI gets built.

Usage: python3 parse_camping_schema.py camping-schema.xlsx > campSchedule.json
"""
import sys
import json
import re
from openpyxl import load_workbook

CAMP_NAME_TO_ID = {
    "Due Laghi": "duelaghi",
    "Nevegal": "nevegal",
    "Campitello di Fassa": "campitello",
    "Caldonazzo": "caldonazzo",
}

def parse_date_tijd(raw):
    """'13/9 < 16:00' -> ('2026-09-13', False); '14/9 ≥ 16:00' -> ('2026-09-14', True);
       '12/9' -> ('2026-09-12', False); 'Vóór 12/9' -> special-cased by caller."""
    m = re.search(r"(\d{1,2})/(\d{1,2})", raw)
    day, month = int(m.group(1)), int(m.group(2))
    iso = f"2026-{month:02d}-{day:02d}"
    pm = "16:00" in raw and ("≥" in raw or ">=" in raw)
    return iso, pm

def clean_value(v):
    """'—' (em dash placeholder for empty) becomes None; everything else passes through as-is."""
    if v is None:
        return None
    v = str(v).strip()
    if v in ("—", "-", ""):
        return None
    return v

def main(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Camping schema"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    schedule = []
    for raw in rows:
        (raw_datum, raw_camp, raw_actief, raw_diner, raw_dagbriefing, raw_ontbijt, raw_start,
         raw_menu, raw_vp1, raw_vp2, raw_baropen, raw_bardicht, raw_stilte,
         raw_rugafgifte, raw_rugophalen, raw_registratie, raw_opmerking) = raw
        if raw_datum is None:
            continue
        camp_id = CAMP_NAME_TO_ID.get(str(raw_camp).strip())
        if not camp_id:
            raise ValueError(f"Unknown campsite name in sheet: {raw_camp!r}")

        active = str(raw_actief).strip().lower() == "ja"

        if str(raw_datum).strip().lower().startswith("vóór") or str(raw_datum).strip().lower().startswith("voor"):
            # the pre-event preview row — anchor it the day before the first real row's date
            first_real_iso, _ = parse_date_tijd(str(rows[1][0]))
            y, m, d = map(int, first_real_iso.split("-"))
            iso = f"{y}-{m:02d}-{d-1:02d}"
            pm = False
        else:
            iso, pm = parse_date_tijd(str(raw_datum))

        row = {
            "iso": iso,
            "pm": pm,
            "camp": camp_id,
            "active": active,
            # Order preserved exactly as the sheet has it: diner, dagbriefing, ontbijt, start.
            "diner": clean_value(raw_diner),
            "dagbriefing": clean_value(raw_dagbriefing),
            "ontbijt": clean_value(raw_ontbijt),
            "start": clean_value(raw_start),
            "menu": clean_value(raw_menu),
            "verzorgingspost1": clean_value(raw_vp1),
            "verzorgingspost2": clean_value(raw_vp2),
            "barOpen": clean_value(raw_baropen),
            "barDicht": clean_value(raw_bardicht),
            "stilte": clean_value(raw_stilte),
            "rugzakjesAfgifte": clean_value(raw_rugafgifte),
            "rugzakjesOphalen": clean_value(raw_rugophalen),
            "registratie": clean_value(raw_registratie),
            "opmerking": clean_value(raw_opmerking),
        }
        # drop keys that are None everywhere for this row's "extra" fields, keep dict small/clean
        schedule.append(row)

    return schedule

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 parse_camping_schema.py camping-schema.xlsx", file=sys.stderr)
        sys.exit(1)
    result = main(sys.argv[1])
    print(json.dumps(result, indent=1, ensure_ascii=False))
